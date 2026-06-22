"""Parse .xlsx bytes from the external SOS export into structured rows.

Column mapping (from the external "Gestión Reclamos y Reintegros" export):

    | Excel Column       | Domain Field    |
    |--------------------|-----------------|
    | N° Gestión         | gestion         |
    | Fecha              | created_at      |
    | Cliente            | claimer_name    |
    | Póliza             | policy_number   |
    | Dominio            | plate           |
    | Tipo               | category        |
    | Motivo             | reason          |
    | Estado             | status          |
    | Usuario Carga      | load_user       |
    | Usuario Respuesta  | response_user   |
    | ITR                | itr             |

The "N° Caso" column is skipped.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from pydantic import BaseModel


# ── Column names ──────────────────────────────────────────────────────────────

COL_GESTION = "N° Gestión"
COL_FECHA = "Fecha"
COL_ASEGURADO = "Cliente"
COL_POLIZA = "Póliza"
COL_PATENTE = "Dominio"
COL_CATEGORIA = "Tipo"
COL_MOTIVO = "Motivo"
COL_ESTADO = "Estado"
COL_CARGA = "Usuario Carga"
COL_RESPONDE = "Usuario Respuesta"
COL_ITR = "ITR"

# Columns that must exist in the Excel header
REQUIRED_COLUMNS = [COL_GESTION]

# All mapped columns (header_name → ParsedRow field)
COLUMN_MAP: dict[str, str] = {
    COL_GESTION: "gestion",
    COL_FECHA: "created_at",
    COL_ASEGURADO: "claimer_name",
    COL_POLIZA: "policy_number",
    COL_PATENTE: "plate",
    COL_CATEGORIA: "category",
    COL_MOTIVO: "reason",
    COL_ESTADO: "status",
    COL_CARGA: "load_user",
    COL_RESPONDE: "response_user",
    COL_ITR: "itr",
}


# ── DTOs ──────────────────────────────────────────────────────────────────────


class ParsedRow(BaseModel):
    """A single parsed row from the Excel file."""

    gestion: int
    created_at: date | None = None
    claimer_name: str = ""
    policy_number: str = ""
    plate: str = ""
    category: str = ""
    reason: str = ""
    status: str = ""
    load_user: str = ""
    response_user: str = ""
    itr: int = 0


class ExcelParseError(Exception):
    """Error information for a row that could not be parsed."""

    def __init__(self, row: int, message: str) -> None:
        self.row = row
        self.message = message
        super().__init__(f"Fila {row}: {message}")


# ── Parser ────────────────────────────────────────────────────────────────────


def parse_excel(
    content: bytes,
    sheet_name: str = "Reclamos y Reintegros",
) -> list[ParsedRow]:
    """Parse ``.xlsx`` bytes into structured rows.

    Args:
        content: Raw bytes of the ``.xlsx`` file.
        sheet_name: Name of the sheet to read.

    Returns:
        List of :class:`ParsedRow` instances.

    Raises:
        ValueError: If the sheet is not found, required columns are missing,
            or the file is empty / unreadable.
    """
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}") from exc

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"La hoja '{sheet_name}' no existe en el archivo. "
            f"Hojas disponibles: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    rows_iter = iter(ws.rows)

    # Read header row
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("El archivo Excel está vacío.")

    # Build column index mapping from header names
    col_indices: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell.value is not None:
            raw = str(cell.value).strip()
            if raw in COLUMN_MAP:
                col_indices[raw] = i

    # Verify required columns
    for req in REQUIRED_COLUMNS:
        if req not in col_indices:
            wb.close()
            raise ValueError(
                f"Columna requerida '{req}' no encontrada en el archivo. "
                f"Columnas encontradas: {list(col_indices.keys())}"
            )

    # Parse data rows
    result: list[ParsedRow] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        try:
            parsed = _parse_row(row, col_indices, row_idx)
            if parsed is not None:
                result.append(parsed)
        except ExcelParseError:
            # Skip rows that cannot be parsed (e.g. non-integer gestion)
            continue

    wb.close()
    return result


# ── Helpers ───────────────────────────────────────────────────────────────────


def _cell_value(row, idx: int):
    """Safely get the value of a cell at *idx* within *row*."""
    if 0 <= idx < len(row):
        cell = row[idx]
        return cell.value if cell is not None else None
    return None


def _str_from(row, col_indices: dict[str, int], name: str) -> str:
    """Get a cell value as a trimmed string (empty string if missing)."""
    idx = col_indices.get(name, -1)
    val = _cell_value(row, idx)
    return str(val).strip() if val is not None else ""


def _int_from(row, col_indices: dict[str, int], name: str) -> int:
    """Get a cell value as an integer (0 if missing or unparseable)."""
    idx = col_indices.get(name, -1)
    val = _cell_value(row, idx)
    if val is not None:
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            pass
    return 0


def _date_from(row, col_indices: dict[str, int], name: str) -> date | None:
    """Get a cell value as a date (``None`` if missing or unparseable).

    Handles ``datetime``, ``date``, and ``"DD/MM/YYYY"`` string values.
    """
    idx = col_indices.get(name, -1)
    val = _cell_value(row, idx)
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        stripped = val.strip()
        try:
            return datetime.strptime(stripped, "%d/%m/%Y").date()
        except ValueError:
            pass
    return None


def _parse_row(row, col_indices: dict[str, int], row_idx: int) -> ParsedRow | None:
    """Parse a single data row into a :class:`ParsedRow`, or skip empty rows."""

    # Gestion is required — skip rows without it
    gestion_raw = _cell_value(row, col_indices[COL_GESTION])
    if gestion_raw is None:
        return None

    try:
        gestion = int(float(str(gestion_raw).strip()))
    except (ValueError, TypeError):
        raise ExcelParseError(
            row=row_idx,
            message=f"'{gestion_raw}' no es un número de gestión válido.",
        )

    return ParsedRow(
        gestion=gestion,
        created_at=_date_from(row, col_indices, COL_FECHA),
        claimer_name=_str_from(row, col_indices, COL_ASEGURADO),
        policy_number=_str_from(row, col_indices, COL_POLIZA),
        plate=_str_from(row, col_indices, COL_PATENTE),
        category=_str_from(row, col_indices, COL_CATEGORIA),
        reason=_str_from(row, col_indices, COL_MOTIVO),
        status=_str_from(row, col_indices, COL_ESTADO),
        load_user=_str_from(row, col_indices, COL_CARGA),
        response_user=_str_from(row, col_indices, COL_RESPONDE),
        itr=_int_from(row, col_indices, COL_ITR),
    )
